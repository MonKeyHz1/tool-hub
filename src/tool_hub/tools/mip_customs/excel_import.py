"""Excel 批量导入引擎 - 从 Excel 文件中读取数据并提交到 MIP API。

核心流程：
1. 读取 Excel 文件，解析表头和行数据
2. 按运单号分组（一个运单号可能对应多行商品）
3. 为每个分组构建 CustomsItem 模型
4. 并发提交到 MIP API（创建或更新模式）
5. 返回导入结果汇总
"""

from __future__ import annotations

import asyncio
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import structlog
from tenacity import retry, retry_if_exception, stop_after_attempt, wait_exponential

from .client import MIPAsyncClient
from .exceptions import MIPAPIError, MIPAuthError, MIPDuplicateError, MIPValidationError
from .import_config import ImportDefaults
from .import_state import ImportState
from .import_task import ImportTaskManager, ProgressCallback
from .models import CustomsItem, GoodsItem, Order, Receiver, Sender

logger = structlog.get_logger()


async def _noop_progress(progress: int, total: int, message: str) -> None:
    """默认进度回调（什么都不做）。"""


# ======================================================================
# 数据结构
# ======================================================================


@dataclass
class RowError:
    """单行导入错误详情。"""

    row_number: int
    tracking_number: str
    error_type: str   # VALIDATION / DUPLICATE / API_ERROR
    message: str


@dataclass
class ImportBatch:
    """批量导入结果汇总。"""

    source_file: str
    total_rows: int = 0
    processed_rows: int = 0
    success_count: int = 0
    failure_count: int = 0
    started_at: datetime | None = None
    completed_at: datetime | None = None
    errors: list[RowError] = field(default_factory=list)

    def __str__(self) -> str:
        duration = ""
        if self.started_at and self.completed_at:
            delta = self.completed_at - self.started_at
            duration = f" in {delta.total_seconds():.1f}s"
        return (
            f"ImportBatch({self.source_file}): "
            f"{self.success_count}/{self.processed_rows} succeeded{duration}"
        )


# ======================================================================
# Excel 解析
# ======================================================================

# 必需的列名（Excel 第二行英文表头）
_REQUIRED_COLUMNS = {
    "TRACKING_NUMBER/MAIL_ID",
    "GOODS_NM",
    "QTY",
    "CONSIGNEE_NM1",
}

# 列名 -> 类型转换器
_COLUMN_PARSERS: dict[str, type] = {
    "TRACKING_NUMBER/MAIL_ID": str,
    "MAIL_BAG_NUMBER": str,
    "GOODS_NM": str,
    "QTY": int,
    "ORDER_QTY": int,
    "QTY_UNIT": str,
    "HSCODE": str,
    "CONSIGNEE_NM1": str,
    "CONSIGNEE_NM2": str,
    "CONSIGNEE_CNTRY_CD": str,
    "CONSIGNEE_CNTRY_NM": str,
    "CONSIGNEE_NATINALITY": str,
    "CONSIGNEE_ADDR": str,
    "CONSIGNEE_ADDR1": str,
    "CONSIGNEE_ADDR2": str,
    "CONSIGNEE_ADDR3": str,
    "CONSIGNEE_TEL": str,
    "CONSIGNEE_ZIPCODE": str,
    "WGT": Decimal,
    "NET_WGT": Decimal,
}


def _parse_value(raw: Any, parser: type) -> Any:
    """将 Excel 单元格原始值转换为目标 Python 类型。

    Args:
        raw: Excel 单元格原始值。
        parser: 目标类型（str / int / Decimal）。

    Returns:
        转换后的值，None 则返回 None。
    """
    if raw is None:
        return None
    if parser is str:
        return str(raw).strip()
    if parser is int:
        return int(raw)
    if parser is Decimal:
        if isinstance(raw, Decimal):
            return raw
        return Decimal(str(raw))
    return raw


def _read_excel_headers(ws: Any) -> dict[str, int]:
    """读取 Excel 第二行（英文表头），返回 表头名 -> 列索引 的映射。

    处理重复表头：第二个 QTY 列会被重命名为 ORDER_QTY。
    """
    headers: dict[str, int] = {}
    seen_counts: dict[str, int] = {}
    for idx, cell in enumerate(ws[2], start=0):
        if cell.value:
            raw = str(cell.value).strip()
            seen_counts[raw] = seen_counts.get(raw, 0) + 1
            if seen_counts[raw] > 1:
                # 重复列自动去重
                key = "ORDER_QTY" if raw == "QTY" else f"{raw}_{seen_counts[raw]}"
            else:
                key = raw
            headers[key] = idx
    return headers


def _validate_headers(headers: dict[str, int]) -> None:
    """校验 Excel 表头是否包含所有必需的列。

    Raises:
        ValueError: 缺少必需列时抛出。
    """
    missing = set()
    for req in _REQUIRED_COLUMNS:
        if req == "QTY":
            if "QTY" not in headers and "ORDER_QTY" not in headers:
                missing.add("QTY")
        elif req not in headers:
            missing.add(req)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")


def _read_row_data(
    ws: Any,
    headers: dict[str, int],
    min_row: int = 3,
) -> list[dict[str, Any]]:
    """读取所有数据行（从第 min_row 行开始），返回 dict 列表。

    跳过完全为空的行。
    """
    rows: list[dict[str, Any]] = []
    for excel_row in ws.iter_rows(min_row=min_row, values_only=False):
        if all(cell.value is None for cell in excel_row):
            continue
        row: dict[str, Any] = {}
        for hdr, idx in headers.items():
            parser = _COLUMN_PARSERS.get(hdr, str)
            try:
                row[hdr] = _parse_value(excel_row[idx].value, parser)
            except (ValueError, TypeError, InvalidOperation):
                row[hdr] = None
        rows.append(row)
    return rows


def _group_rows_by_tracking(
    rows: list[dict[str, Any]],
) -> dict[str, list[tuple[int, dict[str, Any]]]]:
    """按 TRACKING_NUMBER/MAIL_ID 分组，保留 Excel 行号。

    返回 {运单号: [(行号, 行数据), ...]}
    """
    groups: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for row_num, row in enumerate(rows, start=3):
        tracking = row.get("TRACKING_NUMBER/MAIL_ID")
        if not tracking:
            continue
        groups.setdefault(tracking, []).append((row_num, row))
    return groups


# ======================================================================
# 模型构建器
# ======================================================================


def _build_receiver(rows: list[dict[str, Any]]) -> Receiver:
    """从 Excel 行构建收货人信息。"""
    first = rows[0]
    return Receiver(
        country_code=first.get("CONSIGNEE_CNTRY_CD", ""),
        country_name=first.get("CONSIGNEE_CNTRY_NM", ""),
        nationality=first.get("CONSIGNEE_NATINALITY"),
        name_1=first.get("CONSIGNEE_NM1", ""),
        name_2=first.get("CONSIGNEE_NM2") or "",
        address=first.get("CONSIGNEE_ADDR", ""),
        address_1=first.get("CONSIGNEE_ADDR1"),
        address_2=first.get("CONSIGNEE_ADDR2"),
        address_3=first.get("CONSIGNEE_ADDR3"),
        zipcode=first.get("CONSIGNEE_ZIPCODE"),
        telephone=(
            str(first.get("CONSIGNEE_TEL", "")) if first.get("CONSIGNEE_TEL") else None
        ),
    )


def _build_order(
    tracking: str, rows: list[dict[str, Any]], defaults: ImportDefaults
) -> Order:
    """从 Excel 行和默认值构建运单信息。"""
    first = rows[0]
    order_qty = first.get("ORDER_QTY") or first.get("QTY")
    return Order(
        order_id=first.get("ORDER_ID") or defaults.generate_order_id(tracking),
        mail_bag_number=first.get("MAIL_BAG_NUMBER"),
        weight=f"{first.get('WGT')}" if first.get("WGT") else None,
        net_weight=f"{first.get('NET_WGT')}" if first.get("NET_WGT") else None,
        weight_unit=first.get("WGT_UNIT", "kg"),
        quantity=order_qty if isinstance(order_qty, int) else 1,
        quantity_unit=first.get("QTY_UNIT"),
    )


def _build_goods_items(
    rows: list[dict[str, Any]], defaults: ImportDefaults
) -> list[GoodsItem]:
    """从 Excel 行构建商品列表（每行一个商品）。"""
    items: list[GoodsItem] = []
    for row in rows:
        qty = row.get("QTY")
        if not isinstance(qty, int) or qty < 1:
            continue
        items.append(
            GoodsItem(
                goods_name=row.get("GOODS_NM", ""),
                hs_code=str(row.get("HSCODE")) if row.get("HSCODE") else None,
                quantity=qty,
                quantity_unit=row.get("QTY_UNIT", ""),
                price_1=f"{defaults.price}" if not row.get("PRICE1") else f"{row['PRICE1']}",
                price_curr_1=defaults.price_currency,
            )
        )
    return items


def _build_sender(rows: list[dict[str, Any]], defaults: ImportDefaults) -> Sender:
    """从 Excel 行和默认值构建发货人信息。"""
    row = rows[0]
    return Sender(
        country_code=(
            str(row.get("SHIPPER_CNTRY_CD"))
            if row.get("SHIPPER_CNTRY_CD")
            else defaults.sender_country_code
        ),
        country_name=(
            str(row.get("SHIPPER_CNTRY_NM"))
            if row.get("SHIPPER_CNTRY_NM")
            else defaults.sender_country_name
        ),
        nationality=(
            str(row.get("SHIPPER_NATINALITY"))
            if row.get("SHIPPER_NATINALITY")
            else defaults.sender_nationality or None
        ),
        name_1=(
            str(row.get("SHIPPER_NM1"))
            if row.get("SHIPPER_NM1")
            else defaults.sender_name_1
        ),
        name_2=(
            str(row.get("SHIPPER_NM2"))
            if row.get("SHIPPER_NM2")
            else defaults.sender_name_2 or ""
        ),
        registration=(
            str(row.get("SHIPPER_REG"))
            if row.get("SHIPPER_REG")
            else defaults.sender_registration or None
        ),
        address=(
            str(row.get("SHIPPER_ADDR"))
            if row.get("SHIPPER_ADDR")
            else defaults.sender_address
        ),
        address_1=(
            str(row.get("SHIPPER_ADDR1"))
            if row.get("SHIPPER_ADDR1")
            else defaults.sender_address_1 or None
        ),
        address_2=(
            str(row.get("SHIPPER_ADDR2"))
            if row.get("SHIPPER_ADDR2")
            else defaults.sender_address_2 or None
        ),
        address_3=(
            str(row.get("SHIPPER_ADDR3"))
            if row.get("SHIPPER_ADDR3")
            else defaults.sender_address_3 or None
        ),
        zipcode=(
            str(row.get("SHIPPER_ZIPCODE"))
            if row.get("SHIPPER_ZIPCODE")
            else defaults.sender_zipcode or None
        ),
        telephone=(
            str(row.get("SHIPPER_TEL"))
            if row.get("SHIPPER_TEL")
            else defaults.sender_telephone or None
        ),
    )


def _build_customs_item(
    tracking: str,
    row_tuples: list[tuple[int, dict[str, Any]]],
    defaults: ImportDefaults,
) -> CustomsItem:
    """从运单号分组中构建一个 CustomsItem。

    Args:
        tracking: 运单号。
        row_tuples: [(行号, 行数据), ...]。
        defaults: 导入默认值。

    Returns:
        构建好的 CustomsItem 模型。

    Raises:
        ValueError: 如果没有有效的商品数据。
    """
    rows = [r for _, r in row_tuples]
    goods = _build_goods_items(rows, defaults)
    if not goods:
        raise ValueError(f"No valid goods items for tracking {tracking}")
    record = rows[0]
    return CustomsItem(
        tracking_number=tracking,
        mail_id=tracking,
        status_id=str(record.get("STATUS_ID")) if record.get("STATUS_ID") else "1",
        status_cd=(
            str(record.get("STATUS_CD"))
            if record.get("STATUS_CD")
            else "CUSTOMS_ORDER_PLACED"
        ),
        order=_build_order(tracking, rows, defaults),
        goods=goods,
        sender=_build_sender(rows, defaults),
        receiver=_build_receiver(rows),
        bl_no=str(record.get("BL_NO")) if record.get("BL_NO") else None,
        transport_type=(
            str(record.get("TRANSPORT_TYPE"))
            if record.get("TRANSPORT_TYPE")
            else defaults.transport_type
        ),
        trans_fare=(
            str(record.get("TRANS_FARE"))
            if record.get("TRANS_FARE")
            else defaults.trans_fare
        ),
        trans_fare_curr=(
            str(record.get("TRANS_FARE_CURR"))
            if record.get("TRANS_FARE_CURR")
            else defaults.trans_fare_curr
        ),
        is_diplomat=record.get("IS_DIPLOMAT") or defaults.is_diplomat,
        customer_id=(
            str(record.get("CUSTOMERID"))
            if record.get("CUSTOMERID")
            else defaults.customer_id or ""
        ),
        customer_name=(
            str(record.get("CUSTOMERNAME"))
            if record.get("CUSTOMERNAME")
            else defaults.customer_name or ""
        ),
        company_name=(
            str(record.get("COMP_NAME"))
            if record.get("COMP_NAME")
            else defaults.company_name or ""
        ),
        company_register=(
            str(record.get("COMP_REGISTER"))
            if record.get("COMP_REGISTER")
            else defaults.company_register
        ),
        company_addr=(
            str(record.get("COMP_ADDR"))
            if record.get("COMP_ADDR")
            else defaults.company_addr
        ),
        company_tel=(
            str(record.get("COMP_TEL"))
            if record.get("COMP_TEL")
            else defaults.company_tel
        ),
        ecommerce_type="N",
        ecommerce_link="",
        reg_date=datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S"),
        mail_date=datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S"),
    )


# ======================================================================
# 导入器
# ======================================================================


class ExcelImporter:
    """Excel 批量导入器 - 核心引擎。

    负责从 Excel 文件读取数据、构建模型、并发提交到 MIP API。
    """

    # 最大并发数
    _MAX_CONCURRENCY = 20

    def __init__(
        self,
        client: MIPAsyncClient,
        defaults: ImportDefaults | None = None,
        state: ImportState | None = None,
        use_create_api: bool = False,
        progress_callback: ProgressCallback | None = None,
        task_manager: ImportTaskManager | None = None,
        task_id: str = "",
    ) -> None:
        """初始化导入器。

        Args:
            client: MIP 异步客户端。
            defaults: 导入默认值配置。
            state: 导入状态存储（用于去重）。
            use_create_api: True=创建模式, False=更新模式。
            progress_callback: 进度回调函数 (progress, total, message)。
            task_manager: 任务管理器（用于检查取消状态）。
            task_id: 当前任务 ID。
        """
        self.client = client
        self.defaults = defaults or ImportDefaults()
        self.state = state or ImportState()
        self.use_create_api = use_create_api
        self.progress_callback = progress_callback or _noop_progress
        self.task_manager = task_manager
        self.task_id = task_id
        self._logger = logger.bind(component="ExcelImporter")

    async def _call_api_with_retry(self, item: CustomsItem) -> None:
        """调用 MIP API（创建或更新），支持自动重试。

        只对 5xx 错误重试，认证/校验/重复错误不重试。
        """

        def _should_retry(exc: BaseException) -> bool:
            """判断异常是否应该重试。"""
            if not isinstance(exc, MIPAPIError):
                return False
            if isinstance(exc, (MIPAuthError, MIPValidationError, MIPDuplicateError)):
                return False
            status = getattr(exc, "status_code", None)
            return status is not None and status >= 500

        if self.use_create_api:
            @retry(
                retry=retry_if_exception(_should_retry),
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=1, max=10),
                reraise=True,
            )
            async def _call() -> None:
                await self.client.create_item(item)

            await _call()
        else:
            @retry(
                retry=retry_if_exception(_should_retry),
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=1, max=10),
                reraise=True,
            )
            async def _call() -> None:
                await self.client.update_item(item)

            await _call()

    async def _process_group(
        self,
        tracking: str,
        row_tuples: list[tuple[int, dict[str, Any]]],
        path: Path,
        batch: ImportBatch,
        semaphore: asyncio.Semaphore,
        cancel_event: asyncio.Event,
    ) -> None:
        """处理一个运单号分组（含并发控制）。

        Args:
            tracking: 运单号。
            row_tuples: [(行号, 行数据), ...]。
            path: Excel 文件路径。
            batch: 导入结果对象。
            semaphore: 并发信号量。
            cancel_event: 取消事件（认证失败时触发）。
        """
        if cancel_event.is_set():
            return

        if self.task_manager and self.task_id and self.task_manager.is_cancelled(self.task_id):
            cancel_event.set()
            return

        # 跳过已处理的记录
        if self.state.is_processed(tracking):
            batch.success_count += 1
            self._logger.info("skip_already_processed", tracking_number=tracking)
            return

        async with semaphore:
            if cancel_event.is_set():
                return

            log = self._logger.bind(tracking_number=tracking)
            try:
                item = _build_customs_item(tracking, row_tuples, self.defaults)
                await self._call_api_with_retry(item)
                self.state.mark_processed(
                    tracking_number=tracking,
                    source_file=str(path),
                    status="success",
                )
                batch.success_count += 1
                log.info("item_updated", goods_count=len(item.goods))
            except MIPAuthError:
                # 认证失败 -> 取消所有后续任务
                log.error("auth_failed")
                batch.errors.append(
                    RowError(
                        row_number=row_tuples[0][0],
                        tracking_number=tracking,
                        error_type="API_ERROR",
                        message="Authentication failed",
                    )
                )
                batch.failure_count += 1
                cancel_event.set()
                raise
            except MIPValidationError as exc:
                log.warning("validation_failed", error=str(exc))
                batch.errors.append(
                    RowError(
                        row_number=row_tuples[0][0],
                        tracking_number=tracking,
                        error_type="VALIDATION",
                        message=str(exc),
                    )
                )
                batch.failure_count += 1
            except MIPDuplicateError as exc:
                log.warning("duplicate_item", error=str(exc))
                self.state.mark_processed(
                    tracking_number=tracking,
                    source_file=str(path),
                    status="duplicate",
                    error_message=str(exc),
                )
                batch.errors.append(
                    RowError(
                        row_number=row_tuples[0][0],
                        tracking_number=tracking,
                        error_type="DUPLICATE",
                        message=str(exc),
                    )
                )
                batch.failure_count += 1
            except (MIPAPIError, Exception) as exc:
                log.error("item_failed", error=str(exc))
                # 失败行也写入 SQLite，方便重试
                self.state.mark_processed(
                    tracking_number=tracking,
                    source_file=str(path),
                    status="failed",
                    error_message=str(exc),
                )
                batch.errors.append(
                    RowError(
                        row_number=row_tuples[0][0],
                        tracking_number=tracking,
                        error_type="API_ERROR",
                        message=str(exc),
                    )
                )
                batch.failure_count += 1

    async def import_file(
        self, file_path: str | Path, tracking_filter: set[str] | None = None
    ) -> ImportBatch:
        """从 Excel 文件批量导入海关清关数据。

        Args:
            file_path: Excel 文件路径。
            tracking_filter: 可选，只处理指定的运单号列表（用于重试）。

        Returns:
            ImportBatch 包含成功/失败数和错误详情。

        Raises:
            ValueError: Excel 文件缺少必需列。
            MIPAuthError: 认证失败。
        """
        path = Path(file_path)
        log = self._logger.bind(file=str(path))
        log.info("import_start")

        batch = ImportBatch(
            source_file=str(path),
            started_at=datetime.now(UTC),
        )

        try:
            from openpyxl import load_workbook

            await self.progress_callback(0, 0, "正在解析 Excel...")
            log.info("import_start")

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("Excel file has no active worksheet")

            # 读取表头和行数据
            headers = _read_excel_headers(ws)
            _validate_headers(headers)
            rows = _read_row_data(ws, headers)
            batch.total_rows = len(rows)
            log.info("rows_read", total_rows=batch.total_rows)

            # 按运单号分组
            groups = _group_rows_by_tracking(rows)
            # 如果有过滤条件，只保留指定运单号
            if tracking_filter:
                groups = {t: r for t, r in groups.items() if t in tracking_filter}
            batch.processed_rows = sum(len(g) for g in groups.values())
            total_groups = len(groups)
            log.info("groups_formed", group_count=total_groups)

            if total_groups == 0:
                await self.progress_callback(0, 0, "没有需要处理的运单")
                batch.completed_at = datetime.now(UTC)
                return batch

            await self.progress_callback(0, total_groups, "开始导入...")

            # 并发处理
            semaphore = asyncio.Semaphore(self._MAX_CONCURRENCY)
            cancel_event = asyncio.Event()

            completed = 0

            async def _process_with_progress(tracking: str, row_tuples: list[tuple[int, dict[str, Any]]]) -> None:
                nonlocal completed
                await self._process_group(tracking, row_tuples, path, batch, semaphore, cancel_event)
                completed += 1
                # 每处理 5% 或每个都回调，避免过于频繁
                if total_groups <= 50 or completed % max(1, total_groups // 20) == 0 or completed == total_groups:
                    await self.progress_callback(completed, total_groups, f"已处理 {completed}/{total_groups} 单")

            tasks = [
                asyncio.create_task(_process_with_progress(tracking, row_tuples))
                for tracking, row_tuples in groups.items()
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)

            # 检查是否有认证失败导致的中止
            for result in results:
                if isinstance(result, MIPAuthError):
                    raise result

            await self.progress_callback(total_groups, total_groups, "导入完成")

        except MIPAuthError:
            await self.progress_callback(completed, total_groups, "认证失败")
            raise
        except MIPAPIError as exc:
            if getattr(exc, "status_code", None) in (503, 502, 504):
                log.error("service_unavailable", error=str(exc))
                raise
            log.error("api_error", error=str(exc))
            batch.errors.append(
                RowError(
                    row_number=0,
                    tracking_number="",
                    error_type="API_ERROR",
                    message=str(exc),
                )
            )
            batch.failure_count += 1
        except Exception:
            log.exception("import_failed")
            raise
        finally:
            batch.completed_at = datetime.now(UTC)
            log.info(
                "import_complete",
                success=batch.success_count,
                failed=batch.failure_count,
                total=batch.processed_rows,
            )

        return batch
