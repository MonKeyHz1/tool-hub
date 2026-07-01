"""工具中心 - FastAPI 后端服务。

提供 REST API 供前端调用：
- 获取工具列表
- 上传文件
- 执行工具
"""

import uuid
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles

from .config import settings
from .logging import setup_logging
from .models import ExecuteRequest, FileUploadResult
from .tools import registry  # 触发工具自动注册
from .tools.temu_gateway.router import router as temu_gateway_router
from .tools.encoding_converter.router import router as encoding_converter_router
from .tools.pdd_order.router import router as pdd_order_router
from .tools.push_weight.router import router as push_weight_router
from .tools.financial_system.router import router as financial_system_router

# 初始化日志
setup_logging()

# 创建 FastAPI 应用
app = FastAPI(
    title="工具中心",
    description="集成多个业务工具的后端服务，支持 Excel 批量导入等操作",
    version="0.1.0",
)

# 配置 CORS（允许前端跨域访问）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# 挂载 Temu 网关专属路由
app.include_router(temu_gateway_router)
# 挂载 PDD 下单专属路由
app.include_router(pdd_order_router)
# 挂载推送重量路由
app.include_router(push_weight_router)
# 挂载财务系统路由
app.include_router(financial_system_router)
# 挂载编码转换路由
app.include_router(encoding_converter_router)

# 挂载静态文件目录（供编码转换等工具下载文件使用）
app.mount("/files", StaticFiles(directory=str(settings.upload_path)), name="files")


@app.on_event("startup")
async def startup_event() -> None:
    """服务启动时执行。"""
    upload_dir = settings.upload_path
    upload_dir.mkdir(parents=True, exist_ok=True)
    print(f"[工具中心] 已启动，共注册 {registry.tool_count} 个工具")
    print(f"[工具中心] 上传目录: {upload_dir}")
    for t in registry.list_tools():
        print(f"  - {t['tool_id']}: {t['tool_name']}")


@app.get("/api/tools")
async def list_tools() -> dict[str, Any]:
    """获取所有可用工具的列表。"""
    tools = registry.list_tools()
    return {"tools": tools, "count": len(tools)}


@app.get("/api/tools/{tool_id}")
async def get_tool_detail(tool_id: str) -> dict[str, Any]:
    """获取指定工具的详细信息。"""
    tool = registry.get_tool(tool_id)
    if tool is None:
        raise HTTPException(status_code=404, detail=f"工具 '{tool_id}' 不存在")
    meta = tool.get_meta()
    return meta.model_dump()


@app.post("/api/files/upload")
async def upload_file(file: UploadFile = File(...)) -> FileUploadResult:
    """上传文件（供工具使用）。"""
    contents = await file.read()
    if len(contents) > settings.max_upload_size:
        raise HTTPException(
            status_code=413,
            detail=f"文件大小超过限制 ({settings.max_upload_size // 1024 // 1024}MB)",
        )

    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    ext = Path(file.filename).suffix.lower()
    supported_extensions: set[str] = set()
    for t in registry.list_tools():
        supported_extensions.update(t.get("accepted_file_types", []))

    if supported_extensions and ext not in supported_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件类型 '{ext}'，支持的类型: {sorted(supported_extensions)}",
        )

    file_id = uuid.uuid4().hex[:12]
    stored_name = f"{file_id}{ext}"
    stored_path = settings.upload_path / stored_name

    with open(stored_path, "wb") as f:
        f.write(contents)

    from datetime import UTC, datetime

    result = FileUploadResult(
        file_id=file_id,
        original_name=file.filename,
        size=len(contents),
        stored_path=str(stored_path),
        uploaded_at=datetime.now(UTC).isoformat(),
    )
    return result


@app.post("/api/tools/execute")
async def execute_tool(request: Request, exec_request: ExecuteRequest) -> dict[str, Any]:
    """执行指定的工具。"""
    tool = registry.get_tool(exec_request.tool_id)
    if tool is None:
        raise HTTPException(status_code=404, detail=f"工具 '{exec_request.tool_id}' 不存在")

    file_path: Path | None = None
    if exec_request.file_id:
        upload_dir = settings.upload_path
        candidates = list(upload_dir.glob(f"{exec_request.file_id}.*"))
        if candidates:
            file_path = candidates[0]
        else:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "message": f"文件 ID '{exec_request.file_id}' 对应的文件不存在",
                },
            )

    request_host = request.headers.get("host", "")
    result = await tool.execute(params=exec_request.params, file_path=file_path, request_host=request_host)
    return result.model_dump()


@app.get("/api/tool-state/{tool_id}")
async def get_tool_state(tool_id: str) -> dict[str, Any]:
    """获取工具的持久化状态数据（上次保存的表单值）。"""
    from .tool_state import get_state

    return {"tool_id": tool_id, "data": get_state(tool_id)}


@app.post("/api/tool-state/{tool_id}")
async def save_tool_state(tool_id: str, data: dict[str, Any]) -> dict[str, str]:
    """保存工具的输入状态数据（在每次请求前调用）。"""
    from .tool_state import save_state

    save_state(tool_id, data)
    return {"status": "ok"}


@app.get("/api/health")
async def health_check() -> dict[str, str]:
    """健康检查接口。"""
    return {"status": "ok", "version": "0.1.0"}


@app.post("/api/files/cleanup")
async def cleanup_upload_files() -> dict[str, Any]:
    """删除 uploads 目录下所有上传/生成的文件（保留 .gitkeep）。"""
    upload_dir = settings.upload_path
    deleted = 0
    failed = []
    for item in upload_dir.iterdir():
        if item.name == ".gitkeep":
            continue
        try:
            if item.is_file() or item.is_symlink():
                item.unlink()
            elif item.is_dir():
                import shutil
                shutil.rmtree(item)
            deleted += 1
        except Exception as e:
            failed.append({"path": str(item), "error": str(e)})
    return {
        "success": len(failed) == 0,
        "message": f"已删除 {deleted} 个文件/目录" + (f"，{len(failed)} 个失败" if failed else ""),
        "data": {"deleted": deleted, "failed": failed},
    }


@app.post("/api/mip-customs/clear-state")
async def clear_mip_state() -> dict[str, Any]:
    """清除 MIP 海关清关的去重状态记录。"""
    from .tools.mip_customs.import_state import ImportState

    state = ImportState()
    deleted = state.clear()
    return {"success": True, "message": f"已清除 {deleted} 条状态记录", "deleted": deleted}


@app.get("/api/mip-customs/template")
async def download_mip_template():
    """生成并下载 MIP 导入 Excel 模板。"""
    from io import BytesIO

    import openpyxl
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入模板"

    # 中文表头
    cn_headers = [
        "运单号", "邮袋号", "品名", "数量", "订单数量", "单位",
        "HS编码", "收件人1", "收件人2", "国家代码", "国家名称",
        "国籍", "地址", "地址1", "地址2", "地址3", "电话", "邮编",
        "毛重(kg)", "净重(kg)",
    ]
    ws.append(cn_headers)

    # 英文表头（代码实际读取的）
    en_headers = [
        "TRACKING_NUMBER/MAIL_ID", "MAIL_BAG_NUMBER", "GOODS_NM", "QTY", "QTY", "QTY_UNIT",
        "HSCODE", "CONSIGNEE_NM1", "CONSIGNEE_NM2", "CONSIGNEE_CNTRY_CD", "CONSIGNEE_CNTRY_NM",
        "CONSIGNEE_NATINALITY", "CONSIGNEE_ADDR", "CONSIGNEE_ADDR1", "CONSIGNEE_ADDR2",
        "CONSIGNEE_ADDR3", "CONSIGNEE_TEL", "CONSIGNEE_ZIPCODE", "WGT", "NET_WGT",
    ]
    ws.append(en_headers)

    # 示例数据行
    ws.append([
        "RX123456789DE", "", "手表", 1, 1, "PCS",
        "9102.11", "Sarangoo Ganbold", "", "MN", "Mongolia",
        "MNG", "Улаанбаатар хот Баянзүрх дүүрэг", "", "", "",
        "95948917", "45043", 0.5, 0.3,
    ])

    # 设置列宽
    for col in range(1, len(en_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "MIP_导入模板.xlsx"
    encoded = filename.encode("utf-8").hex()
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=MIP_template.xlsx; filename*=UTF-8''{filename}"},
    )


@app.get("/api/mip-customs/status/{task_id}")
async def get_mip_import_status(task_id: str) -> dict[str, Any]:
    """获取 MIP 异步导入任务状态（用于前端轮询进度）。"""
    from .tools.mip_customs.import_task import ImportTaskManager

    manager = ImportTaskManager()
    task = await manager.get_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail=f"任务 '{task_id}' 不存在")
    return task.to_dict()


@app.get("/api/mip-customs/results")
async def get_mip_results() -> dict[str, Any]:
    """获取最近一次导入的每行结果（从 SQLite 读取）。"""
    from .tools.mip_customs.import_state import ImportState

    state = ImportState()
    records = state.list_all()
    data = [
        {
            "tracking_number": r.tracking_number,
            "status": r.status,
            "error_message": r.error_message,
            "created_at": r.created_at.isoformat() if hasattr(r.created_at, "isoformat") else str(r.created_at),
        }
        for r in records
    ]
    return {"success": True, "data": data, "total": len(data)}


@app.post("/api/mip-customs/retry")
async def retry_mip_failed(data: dict[str, Any]) -> dict[str, Any]:
    """重试失败的行（需要 file_id 指向之前上传的 Excel）。"""
    from .tools.mip_customs.import_state import ImportState
    from .tools.mip_customs.config import MIPConfig
    from .tools.mip_customs.client import MIPAsyncClient
    from .tools.mip_customs.excel_import import ExcelImporter
    from .tools.mip_customs.import_config import ImportDefaults

    file_id = data.get("file_id", "")
    if not file_id:
        return {"success": False, "message": "请先上传文件"}
    candidates = list(settings.upload_path.glob(f"{file_id}.*"))
    if not candidates:
        return {"success": False, "message": "文件不存在，请重新上传"}

    state = ImportState()
    failed = state.get_failed()
    if not failed:
        return {"success": True, "message": "没有需要重试的记录", "data": {"retried": 0}}

    tracking_filter = {r.tracking_number for r in failed}

    config = MIPConfig()
    token_env = data.get("token_env", "test")
    if token_env == "prod":
        config.access_token = config.access_token_prod
    else:
        config.access_token = config.access_token_test or config.access_token

    try:
        async with MIPAsyncClient(config) as client:
            importer = ExcelImporter(
                client=client,
                defaults=ImportDefaults(),
                state=state,
                use_create_api=data.get("use_create", True),
            )
            batch = await importer.import_file(candidates[0], tracking_filter=tracking_filter)

        return {
            "success": batch.failure_count == 0,
            "message": f"重试完成: {batch.success_count}/{batch.processed_rows} 成功",
            "data": {
                "processed_rows": batch.processed_rows,
                "success_count": batch.success_count,
                "failure_count": batch.failure_count,
            },
            "errors": [
                {"tracking_number": e.tracking_number, "message": e.message}
                for e in batch.errors
            ],
        }
    except Exception as e:
        return {"success": False, "message": str(e)}
